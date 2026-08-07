"""Generic issuer fundamental store with point-in-time provenance."""

from __future__ import annotations

import hashlib
import json
import re
from datetime import UTC, date, datetime, time, timedelta
from html.parser import HTMLParser
from io import BytesIO

import openpyxl
import requests

VERSION = "generic-fundamentals-v1"

DDL = """
CREATE TABLE IF NOT EXISTS issuer_fundamental_documents (
 document_id VARCHAR PRIMARY KEY, issuer VARCHAR, source_name VARCHAR, official_domain VARCHAR,
 source_url VARCHAR, document_url VARCHAR, document_type VARCHAR, reporting_standard VARCHAR,
 reporting_period VARCHAR, publication_date DATE, available_from TIMESTAMPTZ, mime_type VARCHAR,
 source_hash VARCHAR, machine_readable BOOLEAN, parser_status VARCHAR, validation_status VARCHAR,
 manual_review_reason VARCHAR, legacy_document_id VARCHAR, last_checked_at TIMESTAMP);
CREATE TABLE IF NOT EXISTS issuer_fundamental_revisions (
 issuer VARCHAR, metric_id VARCHAR, reporting_period VARCHAR, reporting_standard VARCHAR,
 revision VARCHAR, previous_revision VARCHAR, document_id VARCHAR, recorded_at TIMESTAMP,
 PRIMARY KEY(issuer,metric_id,reporting_period,reporting_standard,revision));
CREATE TABLE IF NOT EXISTS issuer_fundamental_quality_issues (
 issue_id VARCHAR PRIMARY KEY, issuer VARCHAR, document_id VARCHAR, metric_id VARCHAR,
 issue_type VARCHAR, severity VARCHAR, description VARCHAR, status VARCHAR, detected_at TIMESTAMP);
CREATE TABLE IF NOT EXISTS issuer_fundamental_coverage (
 issuer VARCHAR PRIMARY KEY, documents_found INTEGER, structured_documents INTEGER,
 parsed_documents INTEGER, validated_values INTEGER, manual_review_values INTEGER,
 earliest_period DATE, latest_period DATE, latest_publication_date DATE, data_age_days INTEGER,
 metric_families JSON, confidence VARCHAR, status VARCHAR, updated_at TIMESTAMP);
CREATE TABLE IF NOT EXISTS issuer_share_class_rules (
 issuer VARCHAR, secid VARCHAR, share_class VARCHAR, rights VARCHAR, dividend_parity VARCHAR,
 market_spread_status VARCHAR, liquidity_difference VARCHAR, source_url VARCHAR,
 validation_status VARCHAR, PRIMARY KEY(issuer,secid));
"""

VALUE_COLUMNS = {
    "document_id": "VARCHAR",
    "source_hash": "VARCHAR",
    "structured_field": "VARCHAR",
    "parser_version": "VARCHAR",
    "issuer": "VARCHAR",
    "raw_unit": "VARCHAR",
}


class _TextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []

    def handle_data(self, data):
        if data.strip():
            self.parts.append(" ".join(data.split()))


OFFICIAL_HTML = {
    "MOEX": {
        "url": "https://www.moex.com/n98155?nt=200",
        "standard": "IFRS",
        "period": date(2025, 12, 31),
        "published": date(2026, 3, 5),
        "metrics": [
            ("operating_income", "Operating income", "129,041.2", 129041.2, "RUB million", 1e6),
            ("fee_commission_income", "Fee and commission income", "78,655.3", 78655.3, "RUB million", 1e6),
            (
                "net_interest_income",
                "Net interest and other finance income",
                "50,009.8",
                50009.8,
                "RUB million",
                1e6,
            ),
            ("operating_expenses", "Operating expenses", "51,972.6", 51972.6, "RUB million", 1e6),
            (
                "profit_before_other_expenses_tax",
                "Profit before other operating expenses and tax",
                "77,068.6",
                77068.6,
                "RUB million",
                1e6,
            ),
            ("net_profit", "Net profit", "59.4", 59.4, "RUB billion", 1e9),
        ],
    },
    "MTSS": {
        "url": "https://ir.mts.ru/en/news_and_events/corporate_releases/details/742311",
        "standard": "IFRS",
        "period": date(2025, 12, 31),
        "published": date(2026, 3, 5),
        "metrics": [
            ("revenue_q4", "Consolidated Group Revenue", "222.5", 222.5, "RUB billion", 1e9),
            ("oibda_q4", "Group OIBDA", "71.8", 71.8, "RUB billion", 1e9),
            ("net_profit_q4", "net profit", "21.5", 21.5, "RUB billion", 1e9),
            ("net_debt", "net debt", "458.3", 458.3, "RUB billion", 1e9),
            ("net_debt_oibda", "Net debt/LTM OIBDA", "1.6", 1.6, "ratio", 1.0),
        ],
    },
    "PHOR": {
        "url": "https://www.phosagro.com/press/company/phosagro-reports-operating-and-financial-results-for-9m-2025/",
        "standard": "IFRS",
        "period": date(2025, 9, 30),
        "published": date(2025, 11, 20),
        "metrics": [
            ("production", "TOTAL agrochemicals", "9,154.7", 9154.7, "thousand tonnes", 1.0),
            ("sales", "TOTAL agrochemicals", "9,351.0", 9351.0, "thousand tonnes", 1.0),
            ("revenue", "Revenue", "441,736", 441736.0, "RUB million", 1e6),
            ("ebitda", "EBITDA", "145,663", 145663.0, "RUB million", 1e6),
            ("ebitda_margin", "EBITDA margin", "33.0%", 33.0, "percent", 1.0),
            ("net_profit", "Net profit", "95,692", 95692.0, "RUB million", 1e6),
            ("free_cash_flow", "Free cash flow", "59,018", 59018.0, "RUB million", 1e6),
            ("net_debt", "Net debt", "254,522", 254522.0, "RUB million", 1e6),
            ("net_debt_ebitda", "ND/LTM EBITDA", "1.28x", 1.28, "ratio", 1.0),
        ],
    },
}


def ensure_generic_schema(con):
    con.execute(DDL)
    for column, kind in VALUE_COLUMNS.items():
        con.execute(f"ALTER TABLE issuer_fundamental_values ADD COLUMN IF NOT EXISTS {column} {kind}")


def update_coverage(con, issuer: str) -> dict:
    ensure_generic_schema(con)
    row = con.execute(
        """SELECT count(DISTINCT d.document_id),
        count(DISTINCT d.document_id) FILTER(d.machine_readable),
        count(DISTINCT d.document_id) FILTER(d.parser_status='parsed'),
        count(v.metric) FILTER(v.validation_status='validated'),
        count(v.metric) FILTER(v.validation_status='manual_review'),
        min(v.period_end),max(v.period_end),max(d.publication_date),
        list(DISTINCT v.metric) FILTER(v.metric IS NOT NULL)
        FROM issuer_fundamental_documents d LEFT JOIN issuer_fundamental_values v
        ON d.document_id=v.document_id WHERE d.issuer=?""",
        [issuer],
    ).fetchone()
    documents, structured, parsed, validated, review, earliest, latest, publication, metrics = row
    status = (
        "validated_current"
        if validated >= 5
        else "partially_validated"
        if validated
        else "insufficient_official_data"
    )
    confidence = "high" if validated >= 5 else "medium" if validated else "none"
    con.execute(
        """INSERT OR REPLACE INTO issuer_fundamental_coverage VALUES
        (?,?,?,?,?,?,?,?,?,date_diff('day',?,current_date),?,?,?,current_timestamp)""",
        [
            issuer,
            documents,
            structured,
            parsed,
            validated,
            review,
            earliest,
            latest,
            publication,
            publication,
            json.dumps(metrics or []),
            confidence,
            status,
        ],
    )
    return {"issuer": issuer, "documents": documents, "validated": validated, "status": status}


def migrate_sber(con) -> dict:
    """Link validated legacy SBER observations into the generic layer."""
    ensure_generic_schema(con)
    documents = con.execute(
        """SELECT DISTINCT d.document_id,d.source_url,d.document_type,d.accounting_standard,
        d.period_start,d.period_end,d.publication_date,d.available_from,d.mime_type,d.file_hash,
        d.parser_version,d.validation_status
        FROM fundamental_documents d JOIN fundamental_metric_values v USING(document_id)
        WHERE v.quality_status='validated'"""
    ).fetchall()
    for row in documents:
        doc_id, url, kind, standard, start, end, published, available, mime, digest, parser, status = row
        period = f"{start or ''}/{end or ''}"
        domain = "cbr.ru" if "cbr.ru" in url else "sberbank.com"
        con.execute(
            """INSERT OR REPLACE INTO issuer_fundamental_documents
            (document_id,issuer,source_name,official_domain,source_url,document_url,
             document_type,reporting_standard,reporting_period,publication_date,available_from,
             mime_type,source_hash,machine_readable,parser_status,validation_status,
             manual_review_reason,legacy_document_id,last_checked_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,current_timestamp)""",
            [
                doc_id,
                "SBER",
                domain,
                domain,
                url,
                url,
                kind,
                standard,
                period,
                published,
                available,
                mime,
                digest,
                True,
                "parsed",
                status,
                None,
                doc_id,
            ],
        )
    values = con.execute(
        """SELECT v.document_id,v.metric_id,v.accounting_standard,v.period_start,v.period_end,
        v.publication_date,v.available_from,d.source_url,d.file_hash,v.source_page,v.source_table,
        v.raw_value,v.normalized_value,v.normalized_unit,v.revision_id,d.parser_version
        FROM fundamental_metric_values v JOIN fundamental_documents d USING(document_id)
        WHERE v.quality_status='validated'"""
    ).fetchall()
    for row in values:
        (
            doc_id,
            metric,
            standard,
            start,
            end,
            published,
            available,
            url,
            digest,
            page,
            table,
            raw,
            normalized,
            unit,
            revision,
            parser,
        ) = row
        con.execute(
            """INSERT OR REPLACE INTO issuer_fundamental_values
            (secid,metric,reporting_standard,period_start,period_end,publication_date,available_from,
             source,document,page_table,raw_value,normalized_value,unit,validation_status,revision,
             document_id,source_hash,structured_field,parser_version,issuer)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,'validated',?,?,?,?,?,?)""",
            [
                "SBERP",
                metric,
                standard,
                start,
                end,
                published,
                available,
                url,
                url,
                f"{page or ''}/{table or ''}",
                raw,
                normalized,
                unit,
                revision,
                doc_id,
                digest,
                f"{page or ''}/{table or ''}",
                parser,
                "SBER",
            ],
        )
    con.execute(
        """INSERT OR REPLACE INTO issuer_share_class_rules VALUES
        ('SBER','SBERP','preferred','Preferred share; group fundamentals shared with SBER',
        'Dividend parity requires separately validated charter/dividend document',
        'calculated_from_market_prices','lower liquidity than ordinary share',NULL,'partial')"""
    )
    _migrate_sber_derived(con)
    coverage = update_coverage(con, "SBER")
    coverage["values_reused"] = len(values)
    return coverage


def import_official_html(con, issuer: str) -> dict:
    """Validate explicitly mapped values against a hashed official HTML response."""
    ensure_generic_schema(con)
    spec = OFFICIAL_HTML[issuer]
    response = requests.get(spec["url"], headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    response.raise_for_status()
    digest = hashlib.sha256(response.content).hexdigest()
    document_id = digest[:24]
    parser = _TextExtractor()
    parser.feed(response.text)
    text = re.sub(r"\s*,\s*", ",", " ".join(parser.parts))
    available = datetime.combine(spec["published"] + timedelta(days=1), time.min, UTC)
    con.execute(
        """INSERT OR REPLACE INTO issuer_fundamental_documents
        (document_id,issuer,source_name,official_domain,source_url,document_url,
         document_type,reporting_standard,reporting_period,publication_date,available_from,
         mime_type,source_hash,machine_readable,parser_status,validation_status,
         manual_review_reason,legacy_document_id,last_checked_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,current_timestamp)""",
        [
            document_id,
            issuer,
            "official IR",
            spec["url"].split("/")[2],
            spec["url"],
            spec["url"],
            "HTML press release",
            spec["standard"],
            str(spec["period"]),
            spec["published"],
            available,
            response.headers.get("content-type"),
            digest,
            True,
            "parsed",
            "validated",
            None,
            None,
        ],
    )
    inserted = 0
    missing = []
    for metric, label, token, raw, raw_unit, multiplier in spec["metrics"]:
        if label.lower() not in text.lower() or token not in text:
            missing.append(metric)
            continue
        con.execute(
            """INSERT OR REPLACE INTO issuer_fundamental_values
            (secid,metric,reporting_standard,period_start,period_end,publication_date,available_from,
             source,document,page_table,raw_value,normalized_value,unit,validation_status,revision,
             document_id,source_hash,structured_field,parser_version,issuer,raw_unit)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,'validated','original',?,?,?,?,?,?)""",
            [
                issuer,
                metric,
                spec["standard"],
                None,
                spec["period"],
                spec["published"],
                available,
                spec["url"],
                spec["url"],
                f"HTML:{label}",
                raw,
                raw * multiplier,
                "RUB" if "RUB" in raw_unit else raw_unit,
                document_id,
                digest,
                label,
                VERSION,
                issuer,
                raw_unit,
            ],
        )
        inserted += 1
    coverage = update_coverage(con, issuer)
    coverage.update({"inserted": inserted, "missing": missing, "document_hash": digest})
    return coverage


def import_x5_workbook(con) -> dict:
    """Import audited FY2025 rows from X5's official machine-readable data book."""
    ensure_generic_schema(con)
    issuer = "X5"
    url = "https://www.x5.ru/wp-content/uploads/2026/07/financial_and_operating_results_q2_2026.xlsx"
    response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    response.raise_for_status()
    digest = hashlib.sha256(response.content).hexdigest()
    document_id = digest[:24]
    published = date(2026, 7, 16)
    available = datetime.combine(published + timedelta(days=1), time.min, UTC)
    workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True, read_only=True)
    rows = [
        ("revenue", "Profit and Loss", 6, 8, "Выручка", "RUB million", 1e6),
        ("net_profit", "Profit and Loss", 22, 8, "Чистая прибыль за период", "RUB million", 1e6),
        ("adjusted_ebitda", "EBITDA", 14, 8, "Скорр. EBITDA", "RUB million", 1e6),
        ("adjusted_ebitda_margin", "EBITDA", 15, 8, "Рентабельность скорр. EBITDA, %", "ratio", 1.0),
        ("net_debt", "Debt", 11, 8, "Чистый долг до применения МСФО (IFRS) 16", "RUB million", 1e6),
        ("net_debt_ebitda", "Debt", 12, 8, "Чистый долг / EBITDA до применения МСФО (IFRS) 16", "ratio", 1.0),
    ]
    con.execute(
        """INSERT OR REPLACE INTO issuer_fundamental_documents
        (document_id,issuer,source_name,official_domain,source_url,document_url,
         document_type,reporting_standard,reporting_period,publication_date,available_from,
         mime_type,source_hash,machine_readable,parser_status,validation_status,
         manual_review_reason,legacy_document_id,last_checked_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,current_timestamp)""",
        [
            document_id,
            issuer,
            "X5 IR data book",
            "x5.ru",
            url,
            url,
            "XLSX data book",
            "IFRS",
            "FY2025",
            published,
            available,
            response.headers.get("content-type"),
            digest,
            True,
            "parsed",
            "validated",
            None,
            None,
        ],
    )
    inserted = 0
    for metric, sheet, row, column, expected_label, raw_unit, multiplier in rows:
        ws = workbook[sheet]
        label = ws.cell(row=row, column=2).value
        year = ws.cell(row=5, column=column).value
        raw = ws.cell(row=row, column=column).value
        year_value = year.year if isinstance(year, datetime) else year
        if label != expected_label or year_value != 2025 or not isinstance(raw, (int, float)):
            continue
        con.execute(
            """INSERT OR REPLACE INTO issuer_fundamental_values
            (secid,metric,reporting_standard,period_start,period_end,publication_date,available_from,
             source,document,page_table,raw_value,normalized_value,unit,validation_status,revision,
             document_id,source_hash,structured_field,parser_version,issuer,raw_unit)
            VALUES ('X5',?,'IFRS','2025-01-01','2025-12-31',?,?,?,?,?,?,?,?,'validated','original',?,?,?,?,?,?)""",
            [
                metric,
                published,
                available,
                url,
                url,
                f"{sheet}!B{row}:H{row}",
                raw,
                raw * multiplier,
                "RUB" if "RUB" in raw_unit else raw_unit,
                document_id,
                digest,
                f"{sheet}:{expected_label}:2025",
                VERSION,
                issuer,
                raw_unit,
            ],
        )
        inserted += 1
    con.execute(
        """INSERT OR REPLACE INTO issuer_fundamental_quality_issues VALUES
        ('x5-five-transition','X5',?,NULL,'legal_continuity','medium',
        'FIVE/X5 legal-shell continuity must remain document-mapped; no automatic pre-transition merge',
        'manual_review_required',current_timestamp)""",
        [document_id],
    )
    coverage = update_coverage(con, issuer)
    coverage.update({"inserted": inserted, "document_hash": digest})
    return coverage


def _migrate_sber_derived(con):
    """Expose ROE/EPS/BVPS derived from the validated PIT SBER state."""
    state = con.execute(
        """SELECT latest_ras_period,latest_publication_date,roe_ttm,eps_ttm,bvps
        FROM sber_daily_fundamental_state WHERE roe_ttm IS NOT NULL
        ORDER BY trade_date DESC LIMIT 1"""
    ).fetchone()
    if not state:
        return 0
    period, published, roe_value, eps_value, bvps_value = state
    source = con.execute(
        """SELECT document_id,source_url,source_hash,available_from
        FROM issuer_fundamental_documents
        WHERE issuer='SBER' AND reporting_standard='RAS' AND publication_date<=?
        ORDER BY publication_date DESC LIMIT 1""",
        [published],
    ).fetchone()
    if not source:
        return 0
    document_id, url, digest, available = source
    formulas = {
        "roe": (roe_value, "ratio", "net_profit_ttm / latest_equity"),
        "eps": (eps_value, "RUB/share", "net_profit_ttm / ordinary-equivalent shares"),
        "bvps": (bvps_value, "RUB/share", "latest_equity / ordinary-equivalent shares"),
    }
    for metric, (value, unit, formula) in formulas.items():
        con.execute(
            """INSERT OR REPLACE INTO issuer_fundamental_values
            (secid,metric,reporting_standard,period_start,period_end,publication_date,
             available_from,source,document,page_table,raw_value,normalized_value,unit,
             validation_status,revision,document_id,source_hash,structured_field,
             parser_version,issuer,raw_unit)
            VALUES ('SBERP',?,'RAS',NULL,?,?,?,?,?,?,?, ?,?,'validated','derived',
                    ?,?,?,?,'SBER',?)""",
            [
                metric,
                period,
                published,
                available,
                url,
                url,
                formula,
                value,
                value,
                unit,
                document_id,
                digest,
                formula,
                "sber-decision-v5/generic-derived-v1",
                unit,
            ],
        )
    return len(formulas)
